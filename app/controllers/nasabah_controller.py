from flask import Blueprint, render_template, redirect, url_for, flash, request, Response
from flask_login import login_required
from werkzeug.security import generate_password_hash
from .. import db
from ..models.nasabah import Nasabah
from ..models.user import User
from ..forms.nasabah_form import NasabahForm
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime

bp = Blueprint('nasabah', __name__, url_prefix='/nasabah')

@bp.route('/')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    q = request.args.get('q', '')
    nik = request.args.get('nik', '')
    no_telp = request.args.get('no_telp', '')
    min_penghasilan = request.args.get('min_penghasilan', '', type=str)
    max_penghasilan = request.args.get('max_penghasilan', '', type=str)

    query = Nasabah.query

    # Text search
    if q:
        query = query.filter(Nasabah.nama.ilike(f'%{q}%'))

    # NIK filter
    if nik:
        query = query.filter(Nasabah.nik.ilike(f'%{nik}%'))

    # No. Telp filter
    if no_telp:
        query = query.filter(Nasabah.no_telp.ilike(f'%{no_telp}%'))

    # Range filter - penghasilan
    if min_penghasilan:
        try:
            min_val = int(min_penghasilan)
            query = query.filter(Nasabah.penghasilan >= min_val)
        except (ValueError, TypeError):
            pass

    if max_penghasilan:
        try:
            max_val = int(max_penghasilan)
            query = query.filter(Nasabah.penghasilan <= max_val)
        except (ValueError, TypeError):
            pass

    query = query.order_by(Nasabah.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page)
    nasabah = pagination.items

    total = pagination.total

    return render_template('nasabah/nasabah.html', nasabah=nasabah, pagination=pagination, q=q,

                         nik=nik, no_telp=no_telp, min_penghasilan=min_penghasilan, max_penghasilan=max_penghasilan, total=total)

@bp.route('/tambah', methods=['GET', 'POST'])
@login_required
def tambah():
    form = NasabahForm()
    if form.validate_on_submit():
        # Create user account first
        hashed_password = generate_password_hash(form.password.data)
        user = User(username=form.username.data, password=hashed_password, role='nasabah')
        db.session.add(user)
        db.session.flush()  # Get user ID without committing

        # Create nasabah record
        nasabah = Nasabah(
            user_id=user.id,
            nama=form.nama.data,
            nik=form.nik.data,
            alamat=form.alamat.data,
            no_telp=form.no_telp.data,
            pekerjaan=form.pekerjaan.data,
            penghasilan=form.penghasilan.data
        )
        db.session.add(nasabah)
        db.session.commit()
        flash('Nasabah berhasil ditambahkan.', 'success')
        return redirect(url_for('nasabah.index'))

    return render_template('nasabah/nasabah_form.html', form=form, action='Tambah')


@bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id):
    nasabah = Nasabah.query.get_or_404(id)
    form = NasabahForm(obj=nasabah, nasabah_id=id)

    # Populate user choices - only show users with role 'nasabah' that are not linked to any nasabah (or current one)
    available_users = User.query.filter_by(role='nasabah').filter(
        db.or_(
            User.id == nasabah.user_id,  # Include current user if exists
            ~User.id.in_(db.session.query(Nasabah.user_id).filter(Nasabah.user_id.isnot(None)))
        )
    ).all()
    form.user_id.choices = [(0, '-- Pilih User (Opsional) --')] + [(u.id, f"{u.username} ({u.id})") for u in available_users]

    if form.validate_on_submit():
        try:
            # Update nasabah data
            nasabah.nama = form.nama.data
            nasabah.nik = form.nik.data
            nasabah.alamat = form.alamat.data
            nasabah.no_telp = form.no_telp.data
            nasabah.pekerjaan = form.pekerjaan.data
            nasabah.penghasilan = form.penghasilan.data
            nasabah.user_id = form.user_id.data if form.user_id.data != 0 else None

            db.session.commit()
            flash('Nasabah berhasil diperbarui.', 'success')
            return redirect(url_for('nasabah.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui nasabah: {str(e)}', 'danger')

            return render_template('nasabah/nasabah_form.html', form=form, action='Edit', nasabah=nasabah)
    return render_template('nasabah/nasabah_form.html', form=form, action='Edit', nasabah=nasabah)


@bp.route('/detail/<int:id>')
@login_required
def detail(id):
    nasabah = Nasabah.query.get_or_404(id)

    return render_template('nasabah/detail_nasabah.html', nasabah=nasabah)


@bp.route('/hapus/<int:id>', methods=['POST'])
@login_required
def hapus(id):
    nasabah = Nasabah.query.get_or_404(id)
    try:
        db.session.delete(nasabah)
        db.session.commit()
        flash('Nasabah berhasil dihapus.', 'success')
    except SQLAlchemyError as e:
        db.session.rollback()
        # Log can be added here if logging is configured
        flash('Gagal menghapus nasabah. Pastikan tidak ada data terkait yang mencegah penghapusan.', 'danger')
    return redirect(url_for('nasabah.index'))

@bp.route('/export')
@login_required
def export():
    """Export nasabah data to Excel file with proper table formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    q = request.args.get('q', '')
    min_penghasilan = request.args.get('min_penghasilan', '', type=str)
    max_penghasilan = request.args.get('max_penghasilan', '', type=str)

    query = Nasabah.query

    # Apply same filters as index
    if q:
        query = query.filter(Nasabah.nama.ilike(f'%{q}%'))

    if min_penghasilan:
        try:
            min_val = int(min_penghasilan)
            query = query.filter(Nasabah.penghasilan >= min_val)
        except (ValueError, TypeError):
            pass

    if max_penghasilan:
        try:
            max_val = int(max_penghasilan)
            query = query.filter(Nasabah.penghasilan <= max_val)
        except (ValueError, TypeError):
            pass

    nasabah_list = query.order_by(Nasabah.created_at.desc()).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Nasabah"

    # Define styles
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'DATA NASABAH'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Export info
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Diekspor pada: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'
    ws['A2'].alignment = left_align

    # Empty row
    ws.append([])

    # Headers
    headers = ['ID', 'Nama', 'NIK', 'Alamat', 'No. Telepon', 'Pekerjaan', 'Penghasilan (Rp)', 'Tanggal Dibuat']
    ws.append(headers)

    # Apply header styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Data rows
    for n in nasabah_list:
        row_data = [
            n.id,
            n.nama,
            str(n.nik),
            n.alamat,
            n.no_telp,
            n.pekerjaan,
            n.penghasilan,
            n.created_at.strftime('%d-%m-%Y %H:%M')
        ]
        ws.append(row_data)

        # Apply border to data cells
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
            if col_num == 1:  # Center align ID column
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, ws.max_row + 1):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)  # Max width 30

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'data_nasabah_{timestamp}.xlsx'

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )


