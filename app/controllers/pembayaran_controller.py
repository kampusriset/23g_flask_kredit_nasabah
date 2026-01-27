from flask import Blueprint, render_template, redirect, url_for, flash, jsonify, Response, request, current_app
import os
from werkzeug.utils import secure_filename
from flask_login import login_required, current_user
from .. import db
from ..models.pengajuan import Pengajuan, Pembayaran
from datetime import datetime

bp = Blueprint('pembayaran', __name__, url_prefix='/pembayaran')

@bp.route('/pengajuan/<int:pengajuan_id>')
@login_required
def jadwal_pembayaran(pengajuan_id):
    """Menampilkan jadwal pembayaran untuk pengajuan tertentu"""
    if current_user.role == 'nasabah':
        # Untuk nasabah, cari nasabah yang terkait dengan user ini
        from ..models.nasabah import Nasabah
        nasabah = Nasabah.query.filter_by(user_id=current_user.id).first()
        if not nasabah:
            flash('Data nasabah tidak ditemukan. Silakan hubungi admin.', 'warning')
            return redirect(url_for('auth.logout'))

        # Untuk nasabah, tampilkan semua pengajuan yang disetujui miliknya
        pengajuan_list = Pengajuan.query.filter_by(
            nasabah_id=nasabah.id,
            status='disetujui'
        ).all()

        if not pengajuan_list:
            flash('Anda belum memiliki pengajuan yang disetujui.', 'info')
            return redirect(url_for('dashboard.index'))

        # Jika pengajuan_id = 0 atau tidak valid, ambil pengajuan pertama
        if pengajuan_id == 0 or not any(p.id == pengajuan_id for p in pengajuan_list):
            pengajuan_id = pengajuan_list[0].id

        pengajuan = Pengajuan.query.get_or_404(pengajuan_id)
        # Pastikan pengajuan milik nasabah yang login
        if pengajuan.nasabah_id != nasabah.id:
            flash('Akses ditolak.', 'danger')
            return redirect(url_for('dashboard.index'))
    else:
        # Untuk admin, tampilkan pengajuan yang diminta
        pengajuan = Pengajuan.query.get_or_404(pengajuan_id)

    pembayaran_list = Pembayaran.query.filter_by(pengajuan_id=pengajuan_id).order_by(Pembayaran.bulan_ke).all()


    return render_template('pembayaran/jadwal_pembayaran.html', pengajuan=pengajuan, pembayaran_list=pembayaran_list)


@bp.route('/bayar/<int:pembayaran_id>')
@login_required
def bayar(pembayaran_id):
    """Mengarahkan ke pilihan metode pembayaran"""
    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)
    
    # Hanya nasabah yang terkait atau admin yang bisa mengakses
    if current_user.role == 'nasabah':
        if pembayaran.pengajuan.nasabah.user_id != current_user.id:
            flash('Akses ditolak.', 'danger')
            return redirect(url_for('dashboard.index'))
            
    return redirect(url_for('pembayaran.pilih_metode', pembayaran_id=pembayaran_id))

@bp.route('/pilih-metode/<int:pembayaran_id>')
@login_required
def pilih_metode(pembayaran_id):
    """Halaman pilihan metode pembayaran"""
    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)
    return render_template('pembayaran/pilih_metode.html', pembayaran=pembayaran)

@bp.route('/proses/<int:pembayaran_id>/<string:metode>')
@login_required
def proses_pembayaran(pembayaran_id, metode):
    """Halaman instruksi pembayaran berdasarkan metode"""
    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)
    if metode not in ['transfer', 'qris']:
        flash('Metode pembayaran tidak valid.', 'danger')
        return redirect(url_for('pembayaran.pilih_metode', pembayaran_id=pembayaran_id))
    
    return render_template('pembayaran/proses_pembayaran.html', 
                           pembayaran=pembayaran, 
                           metode=metode)

@bp.route('/konfirmasi/<int:pembayaran_id>', methods=['POST'])
@login_required
def konfirmasi_pembayaran(pembayaran_id):
    """Mencatat pembayaran yang sudah dilakukan setelah konfirmasi"""
    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)

    # Handle Upload Bukti Bayar
    if 'bukti_bayar' in request.files:
        file = request.files['bukti_bayar']
        if file and file.filename != '':
            upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'pembayaran')
            os.makedirs(upload_dir, exist_ok=True)

            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"bukti_{pembayaran.id}_{timestamp}_{filename}"
            
            file_path = os.path.join(upload_dir, filename)
            file.save(file_path)
            
            pembayaran.bukti_bayar = f"uploads/pembayaran/{filename}"

    # Simpan status pembayaran
    # Jika ada bukti bayar, status menjadi menunggu verifikasi
    pembayaran.status = 'menunggu_verifikasi'
    pembayaran.tanggal_bayar = datetime.now()

    # Reset denda jika ada (sementara hitung 0 dulu sampai diversifikasi)
    # p.denda = 0  # Denda tetap tersimpan historynya, nanti admin yang validasi final

    db.session.commit()

    flash(f'Bukti pembayaran berhasil diupload. Mohon tunggu verifikasi admin.', 'success')
    return redirect(url_for('pembayaran.jadwal_pembayaran', pengajuan_id=pembayaran.pengajuan_id))

@bp.route('/verifikasi/<int:pembayaran_id>', methods=['POST'])
@login_required
def verifikasi_pembayaran(pembayaran_id):
    """Verifikasi pembayaran oleh admin"""
    if current_user.role != 'admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard.index'))

    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)
    if pembayaran.status != 'menunggu_verifikasi':
        flash('Hanya pembayaran dengan status menunggu verifikasi yang dapat diverifikasi.', 'danger')
        return redirect(url_for('pembayaran.jadwal_pembayaran', pengajuan_id=pembayaran.pengajuan_id))

    pembayaran.status = 'sudah_bayar'
    pembayaran.denda = 0 # Reset denda jika diverifikasi

    # Check if all payments are completed
    pengajuan = pembayaran.pengajuan
    all_payments = Pembayaran.query.filter_by(pengajuan_id=pengajuan.id).all()
    # Note: current payment is already updated in session session but not committed, 
    # so we can check logical state. 
    # But checking DB state directly requires care.
    # Since we set pembayaran.status above, and it shares session, checking objects works.
    
    # Check completeness
    all_lunas = True
    for p in all_payments:
        if p.id == pembayaran.id:
            if p.status != 'sudah_bayar': all_lunas = False # Should be 'sudah_bayar' as set above
        elif p.status != 'sudah_bayar':
            all_lunas = False
            
    if all_lunas:
        pengajuan.status = 'lunas'
        db.session.add(pengajuan)

    db.session.commit()
    flash('Pembayaran berhasil diverifikasi.', 'success')
    return redirect(url_for('pembayaran.jadwal_pembayaran', pengajuan_id=pembayaran.pengajuan_id))

@bp.route('/tolak/<int:pembayaran_id>', methods=['POST'])
@login_required
def tolak_pembayaran(pembayaran_id):
    """Tolak bukti pembayaran oleh admin"""
    if current_user.role != 'admin':
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('dashboard.index'))

    pembayaran = Pembayaran.query.get_or_404(pembayaran_id)
    if pembayaran.status != 'menunggu_verifikasi':
        flash('Hanya pembayaran dengan status menunggu verifikasi yang dapat ditolak.', 'danger')
        return redirect(url_for('pembayaran.jadwal_pembayaran', pengajuan_id=pembayaran.pengajuan_id))

    pembayaran.status = 'belum_bayar'
    # Optional: Delete uploaded file logic here if needed
    pembayaran.bukti_bayar = None # Remove reference

    db.session.commit()
    flash('Bukti pembayaran ditolak. Status kembali menjadi belum bayar.', 'warning')
    return redirect(url_for('pembayaran.jadwal_pembayaran', pengajuan_id=pembayaran.pengajuan_id))

@bp.route('/api/update_denda')
@login_required
def update_denda():
    """API untuk update denda pembayaran yang terlambat"""
    pembayaran_terlambat = Pembayaran.query.filter(
        Pembayaran.status == 'belum_bayar',
        Pembayaran.tanggal_jatuh_tempo < datetime.now().date()
    ).all()

    updated_count = 0
    for p in pembayaran_terlambat:
        old_denda = p.denda
        p.denda = p.denda_terbaru
        if old_denda != p.denda:
            updated_count += 1

    if updated_count > 0:
        db.session.commit()

    return jsonify({
        'success': True,
        'updated_count': updated_count,
        'message': f'Updated denda untuk {updated_count} pembayaran terlambat'
    })

@bp.route('/export/<int:pengajuan_id>')
@login_required
def export_jadwal(pengajuan_id):
    """Export jadwal pembayaran ke Excel file dengan proper table formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    pengajuan = Pengajuan.query.get_or_404(pengajuan_id)
    pembayaran_list = Pembayaran.query.filter_by(pengajuan_id=pengajuan_id).order_by(Pembayaran.bulan_ke).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadwal Pembayaran"

    # Define styles
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = 'JADWAL PEMBAYARAN KREDIT'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Export info
    ws.merge_cells('A2:I2')
    ws['A2'] = f'Diekspor pada: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'
    ws['A2'].alignment = left_align

    # Nasabah info
    ws.merge_cells('A3:I3')
    ws['A3'] = f'Nasabah: {pengajuan.nasabah.nama} | Jumlah Pinjaman: Rp {pengajuan.jumlah_pinjaman:,} | Tenor: {pengajuan.tenor} bulan'
    ws['A3'].alignment = left_align

    # Empty row
    ws.append([])

    # Main headers with merged cells
    ws.merge_cells('B5:I5')
    ws['B5'] = 'JADWAL PEMBAYARAN'
    ws['B5'].font = header_font
    ws['B5'].fill = header_fill
    ws['B5'].alignment = center_align
    ws['B5'].border = border

    # Sub-headers
    sub_headers = [
        'No', 'Bulan Ke', 'Jumlah Angsuran (Rp)', 'Tanggal Jatuh Tempo',
        'Tanggal Pembayaran', 'Status Pembayaran', 'Denda (Rp)', 'Total Bayar (Rp)', 'Keterlambatan (Hari)'
    ]

    ws.append(sub_headers)

    # Apply subheader styling
    for col_num, header in enumerate(sub_headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = border

    # Data rows
    for idx, p in enumerate(pembayaran_list, 1):
        total = p.jumlah_bayar + p.denda

        # Calculate days late if not paid and past due date
        days_late = 0
        if p.status == 'belum_bayar' and p.tanggal_jatuh_tempo < datetime.now().date():
            days_late = (datetime.now().date() - p.tanggal_jatuh_tempo).days

        row_data = [
            idx,  # No
            p.bulan_ke,  # Bulan Ke
            p.jumlah_bayar,  # Jumlah Angsuran (Rp)
            p.tanggal_jatuh_tempo.strftime('%d-%m-%Y'),  # Tanggal Jatuh Tempo
            p.tanggal_bayar.strftime('%d-%m-%Y') if p.tanggal_bayar else '-',  # Tanggal Pembayaran
            p.status.replace('_', ' ').title(),  # Status Pembayaran
            p.denda,  # Denda (Rp)
            total,  # Total Bayar (Rp)
            days_late if days_late > 0 else ''  # Keterlambatan (Hari)
        ]

        ws.append(row_data)

        # Apply border to data cells
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
            if col_num in [1, 2]:  # Center align No and Bulan Ke columns
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-adjust column widths
    for col_num in range(1, len(sub_headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, ws.max_row + 1):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)  # Max width 25

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'jadwal_pembayaran_{pengajuan.nasabah.nama}_{timestamp}.xlsx'

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )


