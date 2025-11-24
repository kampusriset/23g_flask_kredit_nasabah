from flask import Blueprint, render_template, redirect, url_for, flash, request, Response
from flask_login import login_required, current_user
from .. import db, csrf
from ..models.nasabah import Nasabah
from ..models.pengajuan import Pengajuan
from ..models.dokumen import Dokumen

from ..forms.pengajuan_nasabah_form import PengajuanNasabahForm
from ..forms.pengajuan_action_form import PengajuanActionForm
from datetime import datetime, timedelta
import os
from werkzeug.utils import secure_filename
from flask import current_app

bp = Blueprint('pengajuan', __name__, url_prefix='/pengajuan')

@bp.route('/')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    status = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    q = request.args.get('q', '')

    query = Pengajuan.query

    # Filter by user role
    if current_user.role == 'nasabah':
        # For nasabah, only show their own pengajuan
        nasabah = Nasabah.query.filter_by(user_id=current_user.id).first()
        if not nasabah:
            flash('Data nasabah tidak ditemukan.', 'warning')
            return redirect(url_for('dashboard.index'))
        query = query.filter(Pengajuan.nasabah_id == nasabah.id)
    # For admin, show all pengajuan (no additional filter needed)

    # Text search by nama or NIK
    if q:
        query = query.join(Nasabah).filter(
            db.or_(
                Nasabah.nama.ilike(f'%{q}%'),
                Nasabah.nik.ilike(f'%{q}%')
            )
        )

    # Status filter
    if status and status in ['menunggu', 'disetujui', 'ditolak', 'dibatalkan', 'lunas']:
        query = query.filter(Pengajuan.status == status)

    # Date range filter
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Pengajuan.created_at >= date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # Add 1 day to include entire day
            date_to_obj = date_to_obj + timedelta(days=1)
            query = query.filter(Pengajuan.created_at < date_to_obj)
        except ValueError:
            pass

    query = query.order_by(Pengajuan.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page)
    pengajuan = pagination.items

    total = pagination.total

    return render_template('pengajuan/pengajuan.html', pengajuan=pengajuan, pagination=pagination,

                         status=status, date_from=date_from, date_to=date_to, q=q, total=total)



@bp.route('/detail/<int:id>')
@login_required
def detail(id):
    pengajuan = Pengajuan.query.options(db.joinedload(Pengajuan.nasabah)).get_or_404(id)

    return render_template('pengajuan/detail_pengajuan.html', pengajuan=pengajuan)


@bp.route('/setujui/<int:id>', methods=['GET', 'POST'])
@login_required
def setujui(id):
    from ..models.pengajuan import Pembayaran

    pengajuan = Pengajuan.query.options(db.joinedload(Pengajuan.nasabah)).get_or_404(id)
    form = PengajuanActionForm()
    if form.validate_on_submit():
        pengajuan.status = 'disetujui'
        pengajuan.catatan = form.catatan.data
        pengajuan.tanggal_mulai = datetime.now()

        # Generate jadwal pembayaran
        tanggal_mulai = pengajuan.tanggal_mulai.date()
        angsuran = pengajuan.angsuran_per_bulan

        for bulan in range(1, pengajuan.tenor + 1):
            tanggal_jatuh_tempo = tanggal_mulai.replace(day=1) + timedelta(days=32)
            tanggal_jatuh_tempo = tanggal_jatuh_tempo.replace(day=1) - timedelta(days=1)

            pembayaran = Pembayaran(
                pengajuan_id=pengajuan.id,
                bulan_ke=bulan,
                jumlah_bayar=angsuran,
                tanggal_jatuh_tempo=tanggal_jatuh_tempo
            )
            db.session.add(pembayaran)

            # Update tanggal mulai untuk bulan berikutnya
            tanggal_mulai = tanggal_jatuh_tempo + timedelta(days=1)

        # Set tanggal survei otomatis 3 hari setelah persetujuan
        pengajuan.tanggal_survei = datetime.now() + timedelta(days=3)
        pengajuan.status_survei = 'dijadwalkan'

        db.session.commit()
        flash('Pengajuan disetujui, jadwal pembayaran dan survei telah dibuat.', 'success')
        return redirect(url_for('pengajuan.index'))

    return render_template('pengajuan/pengajuan_action.html', form=form, pengajuan=pengajuan, action='Setujui')


@bp.route('/tolak/<int:id>', methods=['GET', 'POST'])
@login_required
def tolak(id):
    pengajuan = Pengajuan.query.options(db.joinedload(Pengajuan.nasabah)).get_or_404(id)
    form = PengajuanActionForm()
    if form.validate_on_submit():
        pengajuan.status = 'ditolak'
        pengajuan.catatan = form.catatan.data
        db.session.commit()
        flash('Pengajuan ditolak.', 'success')
        return redirect(url_for('pengajuan.index'))

    return render_template('pengajuan/pengajuan_action.html', form=form, pengajuan=pengajuan, action='Tolak')


@bp.route('/cancel/<int:id>', methods=['POST'])
@login_required
def cancel(id):
    pengajuan = Pengajuan.query.get_or_404(id)
    if pengajuan.status != 'menunggu':
        flash('Hanya pengajuan dengan status menunggu yang dapat dibatalkan.', 'danger')
        return redirect(url_for('pengajuan.index'))

    # Check if user is nasabah - they can only cancel their own pengajuan
    if current_user.role == 'nasabah':
        nasabah = Nasabah.query.filter_by(user_id=current_user.id).first()
        if not nasabah or pengajuan.nasabah_id != nasabah.id:
            flash('Anda tidak memiliki akses untuk membatalkan pengajuan ini.', 'danger')
            return redirect(url_for('pengajuan.index'))

    pengajuan.status = 'dibatalkan'
    db.session.commit()
    flash('Pengajuan berhasil dibatalkan.', 'success')
    return redirect(url_for('pengajuan.index'))

@bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    pengajuan = Pengajuan.query.get_or_404(id)

    # Check if user is nasabah - they can only delete their own pengajuan
    if current_user.role == 'nasabah':
        from ..models.nasabah import Nasabah
        nasabah = Nasabah.query.filter_by(user_id=current_user.id).first()
        if not nasabah or pengajuan.nasabah_id != nasabah.id:
            flash('Anda tidak memiliki akses untuk menghapus pengajuan ini.', 'danger')
            return redirect(url_for('pengajuan.index'))

    # Allow deletion for rejected, cancelled, approved, or completed applications
    if pengajuan.status not in ['ditolak', 'dibatalkan', 'disetujui', 'lunas']:
        flash('Hanya pengajuan yang ditolak, dibatalkan, disetujui, atau lunas yang dapat dihapus.', 'danger')
        return redirect(url_for('pengajuan.index'))

    # Delete associated dokumen records if any
    Dokumen.query.filter_by(pengajuan_id=id).delete()

    # Delete associated pembayaran records if any
    from ..models.pengajuan import Pembayaran
    Pembayaran.query.filter_by(pengajuan_id=id).delete()

    db.session.delete(pengajuan)
    db.session.commit()
    flash('Pengajuan berhasil dihapus.', 'success')
    return redirect(url_for('pengajuan.index'))

@bp.route('/survei/<int:id>', methods=['GET', 'POST'])
@login_required
@csrf.exempt
def survei(id):
    pengajuan = Pengajuan.query.options(db.joinedload(Pengajuan.nasabah)).get_or_404(id)
    if request.method == 'POST':
        hasil_survei = request.form.get('hasil_survei')
        catatan = request.form.get('catatan_survei', '')

        if hasil_survei == 'lulus':
            pengajuan.status_survei = 'selesai'
            pengajuan.catatan_survei = f"Survei lulus. {catatan}".strip()
        elif hasil_survei == 'tidak_lulus':
            pengajuan.status_survei = 'selesai'
            pengajuan.status = 'ditolak'  # Tolak pengajuan jika survei tidak lulus
            pengajuan.catatan_survei = f"Survei tidak lulus. {catatan}".strip()
        else:
            flash('Pilih hasil survei yang valid.', 'danger')
            return redirect(url_for('pengajuan.survei', id=id))

        db.session.commit()
        flash('Hasil survei berhasil disimpan.', 'success')
        return redirect(url_for('pengajuan.index'))


    return render_template('pengajuan/survei_form.html', pengajuan=pengajuan)


@bp.route('/jadwal-survei')
@login_required
def jadwal_survei():
    """Halaman khusus untuk menampilkan jadwal survei"""
    # Ambil pengajuan yang sudah dijadwalkan survei
    pengajuan_survei = Pengajuan.query.filter(
        Pengajuan.status_survei == 'dijadwalkan'
    ).order_by(Pengajuan.tanggal_survei.asc()).all()

    from datetime import datetime, timedelta
    today = datetime.now().date()
    week_later = today + timedelta(days=7)

    # Hitung jumlah untuk summary cards
    count_today = sum(1 for p in pengajuan_survei if p.tanggal_survei.date() == today)
    count_week = sum(1 for p in pengajuan_survei if today < p.tanggal_survei.date() < week_later)
    count_late = sum(1 for p in pengajuan_survei if p.tanggal_survei.date() < today)

    return render_template('pengajuan/jadwal_survei.html', pengajuan_survei=pengajuan_survei, today=today, week_later=week_later,
                         count_today=count_today, count_week=count_week, count_late=count_late)

@bp.route('/reschedule/<int:id>', methods=['POST'])
@login_required
@csrf.exempt
def reschedule(id):
    """Reschedule jadwal survei sesuai tanggal jatuh tempo pembayaran pertama"""
    pengajuan = Pengajuan.query.get_or_404(id)
    if pengajuan.status_survei != 'dijadwalkan':
        flash('Hanya pengajuan yang dijadwalkan survei yang dapat direschedule.', 'danger')
        return redirect(url_for('pengajuan.jadwal_survei'))

    # Ambil tanggal jatuh tempo pembayaran pertama (terdekat)
    from ..models.pengajuan import Pembayaran
    earliest_due = db.session.query(db.func.min(Pembayaran.tanggal_jatuh_tempo)).filter_by(pengajuan_id=id).scalar()

    if earliest_due:
        pengajuan.tanggal_survei = earliest_due
        db.session.commit()
        flash('Jadwal survei berhasil direschedule sesuai tanggal jatuh tempo.', 'success')
    else:
        flash('Tidak ada jadwal pembayaran untuk direschedule.', 'danger')

    return redirect(url_for('pengajuan.jadwal_survei'))

@bp.route('/ajukan', methods=['GET', 'POST'])
@login_required
def ajukan():
    """Route untuk nasabah mengajukan kredit baru"""
    if current_user.role != 'nasabah':
        flash('Hanya nasabah yang dapat mengajukan kredit.', 'danger')
        return redirect(url_for('dashboard.index'))

    # Cari nasabah yang terkait dengan user ini
    nasabah = Nasabah.query.filter_by(user_id=current_user.id).first()
    if not nasabah:
        flash('Data nasabah tidak ditemukan. Silakan hubungi admin.', 'warning')
        return redirect(url_for('dashboard.index'))

    form = PengajuanNasabahForm()
    if form.validate_on_submit():
        # Cek limit: maksimal 1 pengajuan aktif per nasabah
        existing_active = Pengajuan.query.filter(
            Pengajuan.nasabah_id == nasabah.id,
            Pengajuan.status.in_(['menunggu', 'disetujui'])
        ).first()
        if existing_active:
            flash('Anda sudah memiliki pengajuan aktif. Tidak dapat membuat pengajuan baru.', 'danger')
            return redirect(url_for('dashboard.index'))

        # Cek pembayaran yang belum lunas
        from ..models.pengajuan import Pembayaran
        outstanding_payments = Pembayaran.query.join(Pengajuan).filter(
            Pengajuan.nasabah_id == nasabah.id,
            Pengajuan.status.in_(['disetujui', 'lunas']),
            Pembayaran.status != 'sudah_bayar'
        ).first()
        if outstanding_payments:
            flash('Anda masih memiliki pembayaran yang belum lunas. Harap selesaikan semua pembayaran sebelum mengajukan kredit baru.', 'danger')
            return redirect(url_for('dashboard.index'))

        pengajuan = Pengajuan(
            nasabah_id=nasabah.id,
            jumlah_pinjaman=form.jumlah_pinjaman.data,
            tenor=form.tenor.data,
            tujuan=form.tujuan.data
        )
        db.session.add(pengajuan)
        db.session.commit()

        # Handle KTP upload
        if form.foto_ktp.data:
            # Create upload directory if not exists
            upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'dokumen')
            os.makedirs(upload_dir, exist_ok=True)

            # Secure filename
            filename = secure_filename(form.foto_ktp.data.filename)
            # Add timestamp to avoid conflicts
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"ktp_{pengajuan.id}_{timestamp}_{filename}"

            # Save file
            file_path = os.path.join(upload_dir, filename)
            form.foto_ktp.data.save(file_path)

            # Create dokumen record
            dokumen = Dokumen(
                pengajuan_id=pengajuan.id,
                jenis_dokumen='ktp',
                nama_file=form.foto_ktp.data.filename,  # original filename
                path_file=f"uploads/dokumen/{filename}",  # relative path for web access
                status='sudah_diupload',
                uploaded_by=current_user.id,
                uploaded_at=datetime.utcnow()
            )
            db.session.add(dokumen)
            db.session.commit()

        flash('Pengajuan kredit berhasil diajukan. Silakan tunggu verifikasi dari petugas.', 'success')
        return redirect(url_for('dashboard.index'))


    return render_template('pengajuan/pengajuan_nasabah_form.html', form=form)


@bp.route('/export')
@login_required
def export():
    """Export pengajuan data to Excel file with proper table formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    status = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    q = request.args.get('q', '')

    query = Pengajuan.query

    # Apply same filters as index
    if q:
        query = query.join(Nasabah).filter(
            db.or_(
                Nasabah.nama.ilike(f'%{q}%'),
                Nasabah.nik.ilike(f'%{q}%')
            )
        )

    if status and status in ['menunggu', 'disetujui', 'ditolak', 'dibatalkan', 'lunas']:
        query = query.filter(Pengajuan.status == status)

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Pengajuan.created_at >= date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            query = query.filter(Pengajuan.created_at < date_to_obj)
        except ValueError:
            pass

    pengajuan_list = query.order_by(Pengajuan.created_at.desc()).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pengajuan Kredit"

    # Define styles
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Title
    ws.merge_cells('A1:P1')
    ws['A1'] = 'DATA PENGAJUAN KREDIT'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # Export info
    ws.merge_cells('A2:P2')
    ws['A2'] = f'Diekspor pada: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'
    ws['A2'].alignment = left_align

    # Empty row
    ws.append([])

    # Main headers with merged cells
    ws.merge_cells('B4:G4')
    ws['B4'] = 'INFORMASI NASABAH'
    ws['B4'].font = header_font
    ws['B4'].fill = header_fill
    ws['B4'].alignment = center_align
    ws['B4'].border = border

    ws.merge_cells('H4:K4')
    ws['H4'] = 'INFORMASI PINJAMAN'
    ws['H4'].font = header_font
    ws['H4'].fill = header_fill
    ws['H4'].alignment = center_align
    ws['H4'].border = border

    ws.merge_cells('L4:P4')
    ws['L4'] = 'STATUS & TANGGAL'
    ws['L4'].font = header_font
    ws['L4'].fill = header_fill
    ws['L4'].alignment = center_align
    ws['L4'].border = border

    # Sub-headers
    sub_headers = [
        'No', 'Nama Nasabah', 'NIK', 'No. Telepon', 'Alamat', 'Pekerjaan',
        'Penghasilan (Rp)', 'Jumlah Pinjaman (Rp)', 'Tenor (Bulan)', 'Angsuran per Bulan (Rp)',
        'Tujuan Pinjaman', 'Status Pengajuan', 'Tanggal Pengajuan', 'Tanggal Disetujui',
        'Tanggal Jatuh Tempo Terakhir', 'Catatan'
    ]

    ws.append(sub_headers)

    # Apply subheader styling
    for col_num, header in enumerate(sub_headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = border

    # Data rows
    for idx, p in enumerate(pengajuan_list, 1):
        # Calculate angsuran per bulan if approved
        angsuran = p.angsuran_per_bulan if hasattr(p, 'angsuran_per_bulan') and p.angsuran_per_bulan else (p.jumlah_pinjaman / p.tenor if p.tenor else 0)

        # Calculate last due date if approved
        last_due_date = ''
        if p.status in ['disetujui', 'lunas'] and hasattr(p, 'tanggal_mulai') and p.tanggal_mulai:
            last_due_date = (p.tanggal_mulai + timedelta(days=30 * p.tenor)).strftime('%d-%m-%Y')

        row_data = [
            idx,  # No
            p.nasabah.nama,  # Nama Nasabah
            str(p.nasabah.nik),  # NIK
            p.nasabah.no_telp,  # No. Telepon
            p.nasabah.alamat,  # Alamat
            p.nasabah.pekerjaan,  # Pekerjaan
            p.nasabah.penghasilan,  # Penghasilan
            p.jumlah_pinjaman,  # Jumlah Pinjaman (Rp)
            p.tenor,  # Tenor (Bulan)
            round(angsuran),  # Angsuran per Bulan (Rp)
            p.tujuan,  # Tujuan Pinjaman
            p.status.title(),  # Status Pengajuan
            p.created_at.strftime('%d-%m-%Y %H:%M'),  # Tanggal Pengajuan
            p.tanggal_mulai.strftime('%d-%m-%Y') if hasattr(p, 'tanggal_mulai') and p.tanggal_mulai else '',  # Tanggal Disetujui
            last_due_date,  # Tanggal Jatuh Tempo Terakhir
            p.catatan or ''  # Catatan
        ]

        ws.append(row_data)

        # Apply border to data cells
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
            if col_num == 1:  # Center align No column
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-adjust column widths
    for col_num in range(1, len(sub_headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, ws.max_row + 1):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)  # Max width 30

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'data_pengajuan_kredit_{timestamp}.xlsx'

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )


